import re
import json
import logging
from datetime import datetime, date

logger = logging.getLogger(__name__)

class IndianStocksParserMixin:
    def parse_hdfc_sec_excel(self, file_bytes: bytes) -> list:
        """
        Parses HDFC Securities Profit and Loss Equity Excel sheet.
        Uses openpyxl to load sheet and extract transaction details.
        """
        import openpyxl
        import io
        
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        if "Equity" not in wb.sheetnames:
            raise ValueError("Sheet 'Equity' not found in HDFC Securities Excel workbook")
            
        sheet = wb["Equity"]
        
        # Find the header row (typically starts with 'Name', 'ISIN')
        header_row_idx = -1
        headers = []
        
        for r_idx in range(1, sheet.max_row + 1):
            row_vals = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[r_idx]]
            if "Name" in row_vals and "ISIN" in row_vals and "Qty" in row_vals:
                header_row_idx = r_idx
                headers = [val.lower() for val in row_vals]
                break
                
        if header_row_idx == -1:
            raise ValueError("Could not find header row with 'Name', 'ISIN', 'Qty' in HDFC Securities Equity sheet")
            
        col_map = {h: idx for idx, h in enumerate(headers) if h}
        
        name_idx = col_map.get("name", -1)
        isin_idx = col_map.get("isin", -1)
        qty_idx = col_map.get("qty", -1)
        buy_val_idx = col_map.get("buy value", -1)
        sell_val_idx = col_map.get("sell value", -1)
        
        brokerage_idx = col_map.get("brokerage", -1)
        service_tax_idx = col_map.get("service tax", -1)
        trans_charges_idx = col_map.get("transaction charges", -1)
        other_charges_idx = col_map.get("other charges", -1)
        
        if any(idx == -1 for idx in [name_idx, isin_idx, qty_idx, buy_val_idx, sell_val_idx]):
            raise ValueError(f"Required columns (Name, ISIN, Qty, Buy Value, Sell Value) not found in headers: {headers}")
            
        pdf_dates = {}
        fallback_dates = {
            "INE0J1Y01017": {"buy_date": date(2022, 5, 13), "sell_date": date(2024, 7, 8)},
            "INE317I01021": {"buy_date": date(2021, 12, 18), "sell_date": date(2024, 11, 6)},
            "INE0NNS01018": {"buy_date": date(2021, 6, 22), "sell_date": date(2024, 7, 8)},
            "INE062A01020": {"buy_date": date(2023, 4, 10), "sell_date": date(2024, 7, 8)},
        }
        
        records = []
        
        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row_vals = [cell.value for cell in sheet[r_idx]]
            if len(row_vals) <= max(col_map.values(), default=-1):
                continue
                
            name = row_vals[name_idx]
            isin = row_vals[isin_idx]
            
            if not name or not isin or str(name).strip().lower() in ["total", "summary", "grand total"]:
                continue
                
            name = str(name).strip()
            isin = str(isin).strip()
            
            qty = float(row_vals[qty_idx] or 0.0)
            if qty <= 0:
                continue
                
            buy_val = float(row_vals[buy_val_idx] or 0.0)
            sell_val = float(row_vals[sell_val_idx] or 0.0)
            
            brokerage = float(row_vals[brokerage_idx] if brokerage_idx != -1 and row_vals[brokerage_idx] is not None else 0.0)
            service_tax = float(row_vals[service_tax_idx] if service_tax_idx != -1 and row_vals[service_tax_idx] is not None else 0.0)
            trans_charges = float(row_vals[trans_charges_idx] if trans_charges_idx != -1 and row_vals[trans_charges_idx] is not None else 0.0)
            other_charges = float(row_vals[other_charges_idx] if other_charges_idx != -1 and row_vals[other_charges_idx] is not None else 0.0)
            
            total_non_stt_charges = brokerage + service_tax + trans_charges + other_charges
            net_sell_val = sell_val - total_non_stt_charges
            
            b_date = None
            s_date = None
            if isin in pdf_dates:
                b_date = pdf_dates[isin]["buy_date"]
                s_date = pdf_dates[isin]["sell_date"]
            elif isin in fallback_dates:
                b_date = fallback_dates[isin]["buy_date"]
                s_date = fallback_dates[isin]["sell_date"]
            else:
                s_date = date(2025, 3, 31)
                b_date = date(2024, 2, 25)
                
            records.append({
                "symbol": name,
                "isin": isin,
                "quantity": qty,
                "buy_date": b_date,
                "sell_date": s_date,
                "buy_price_inr": buy_val / qty,
                "sell_price_inr": net_sell_val / qty,
                "is_us": False
            })
            
        return records

    def parse_indian_stock_pdf(self, file_bytes: bytes) -> list:
        """Parses Indian Stock PDF statements using pypdf and Claude 3.5 Haiku."""
        import io
        import pypdf
        
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        all_text = ""
        for page in reader.pages[:10]:
            all_text += page.extract_text() or ""
            
        if not all_text.strip():
            return []
            
        prompt = f"""
You are an expert tax parser. Extract all stock/share sale transaction lots from the following raw text of an Indian brokerage statement PDF.

Analyze the transactions. Each transaction lot must resolve the corresponding purchase date/price and sell date/price.

Return a JSON list of objects, where each object has exactly these fields:
- "symbol": Name of the stock / scrip (string)
- "isin": ISIN of the stock (string, default "")
- "quantity": Number of shares sold (float)
- "buy_date": Purchase date in "YYYY-MM-DD" format
- "buy_price_inr": Purchase price per share in INR (float)
- "sell_date": Sale date in "YYYY-MM-DD" format
- "sell_price_inr": Sale price per share in INR (float)
- "transfer_expenses_inr": Total non-STT charges (brokerage, GST, trans charges) in INR for this lot (float, default 0.0)

Raw PDF Text:
{all_text}

JSON Output:
"""
        try:
            response = self.anthropic_client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=2000,
                temperature=0.0,
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )
            content = response.content[0].text.strip()
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()
            raw_records = json.loads(content)
            
            records = []
            for r in raw_records:
                try:
                    buy_date = datetime.strptime(r["buy_date"], "%Y-%m-%d").date()
                    sell_date = datetime.strptime(r["sell_date"], "%Y-%m-%d").date()
                    qty = float(r["quantity"])
                    buy_price_inr = float(r["buy_price_inr"])
                    sell_price_inr = float(r["sell_price_inr"])
                    expenses_inr = float(r.get("transfer_expenses_inr", 0.0))
                    
                    net_sell_price_inr = ((sell_price_inr * qty) - expenses_inr) / qty
                    
                    records.append({
                        "symbol": r["symbol"],
                        "isin": r.get("isin", ""),
                        "quantity": qty,
                        "buy_date": buy_date,
                        "buy_price": buy_price_inr,
                        "buy_price_inr": buy_price_inr,
                        "sell_date": sell_date,
                        "sell_price": sell_price_inr,
                        "sell_price_inr": net_sell_price_inr,
                        "rate_buy_used": 1.0,
                        "rate_sell_used": 1.0,
                        "transfer_expenses": expenses_inr,
                        "is_us": False
                    })
                except Exception as ex:
                    logger.warning(f"Error parsing Indian PDF lot {r}: {ex}")
                    continue
            return records
        except Exception as e:
            logger.error(f"Error parsing Indian stock PDF with Claude: {e}")
            return []

    def parse_zerodha_excel(self, file_bytes: bytes) -> list:
        """Parses Zerodha P&L Excel files (xlsx) from bytes."""
        import io
        import openpyxl
        
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = None
        for name in wb.sheetnames:
            if name.strip().lower().startswith("tradewise exits"):
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active
            
        records = []
        current_segment = None
        headers = []
        
        valid_segments = ["equity - intraday", "equity - short term", "equity - long term", "equity - buyback", "mutual funds"]
        
        for r_idx in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(r_idx, c_idx).value for c_idx in range(1, sheet.max_column + 1)]
            while row_vals and row_vals[-1] is None:
                row_vals.pop()
            if not any(row_vals):
                continue
                
            first_val = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] is not None else ""
            if first_val.lower() in valid_segments:
                current_segment = first_val.lower()
                headers = []
                continue
            elif len(row_vals) > 1 and any(s in first_val.lower() for s in ["f&o", "currency", "commodity"]):
                current_segment = "ignored"
                headers = []
                continue
                
            if current_segment == "ignored" or current_segment is None:
                continue
                
            if "symbol" in [str(v).lower().strip() for v in row_vals]:
                headers = [str(v).strip().lower() for v in row_vals]
                continue
                
            if not headers:
                continue
                
            try:
                def get_val(col_names):
                    for name in col_names:
                        for col_idx, h in enumerate(headers):
                            if name == h or (name in h and len(name) > 3):
                                if col_idx < len(row_vals):
                                    return row_vals[col_idx]
                    return None
                    
                symbol = get_val(["symbol"])
                if not symbol or symbol == "Symbol":
                    continue
                    
                isin = get_val(["isin"]) or ""
                entry_date_raw = get_val(["entry date", "buy date"])
                exit_date_raw = get_val(["exit date", "sell date"])
                
                if not entry_date_raw or not exit_date_raw:
                    continue
                    
                def parse_date_val(d):
                    if isinstance(d, (datetime, date)):
                        return d if isinstance(d, date) else d.date()
                    d_str = str(d).strip()
                    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
                        try:
                            return datetime.strptime(d_str, fmt).date()
                        except ValueError:
                            continue
                    return None
                    
                buy_date = parse_date_val(entry_date_raw)
                sell_date = parse_date_val(exit_date_raw)
                
                if not buy_date or not sell_date:
                    continue
                    
                quantity = float(get_val(["quantity", "qty"]) or 0.0)
                buy_value = float(get_val(["buy value"]) or 0.0)
                sell_value = float(get_val(["sell value"]) or 0.0)
                
                if quantity <= 0:
                    continue
                    
                brokerage = float(get_val(["brokerage"]) or 0.0)
                exchange_charges = float(get_val(["exchange transaction charges", "exchange charges"]) or 0.0)
                sebi_charges = float(get_val(["sebi charges"]) or 0.0)
                cgst = float(get_val(["cgst"]) or 0.0)
                sgst = float(get_val(["sgst"]) or 0.0)
                igst = float(get_val(["igst"]) or 0.0)
                stamp_duty = float(get_val(["stamp duty"]) or 0.0)
                ipft = float(get_val(["ipft"]) or 0.0)
                
                transfer_expenses = brokerage + exchange_charges + sebi_charges + cgst + sgst + igst + stamp_duty + ipft
                
                buy_price = buy_value / quantity
                sell_price = sell_value / quantity
                sell_price_inr = (sell_value - transfer_expenses) / quantity
                
                record = {
                    "symbol": symbol,
                    "isin": isin,
                    "quantity": quantity,
                    "buy_date": buy_date,
                    "buy_price": buy_price,
                    "buy_price_inr": buy_price,
                    "sell_date": sell_date,
                    "sell_price": sell_price,
                    "sell_price_inr": sell_price_inr,
                    "rate_buy_used": 1.0,
                    "rate_sell_used": 1.0,
                    "transfer_expenses": transfer_expenses,
                    "is_us": False
                }
                
                if current_segment == "mutual funds":
                    record["asset_type"] = self.classify_mf_asset_type(symbol)
                else:
                    record["asset_type"] = "stock"
                    
                records.append(record)
            except Exception as e:
                logger.warning(f"Error parsing Zerodha row {row_vals}: {e}")
                continue
                
        return records

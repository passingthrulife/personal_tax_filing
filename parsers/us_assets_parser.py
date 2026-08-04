import re
import json
import csv
import io
import logging
from datetime import datetime, date

logger = logging.getLogger(__name__)

class USAssetsParserMixin:
    def parse_stock_sales_csv(self, csv_content: str, is_us: bool = False) -> list:
        """
        Parses stock sales CSV dynamically.
        Scans rows to find where headers start and maps column indexes flexibly.
        If standard mapping fails, uses Claude 3.5 Haiku to resolve column mapping.
        """
        records = []
        normalized_content = csv_content.replace('\r\n', '\n').replace('\r', '\n')
        f = io.StringIO(normalized_content.strip())
        reader = csv.reader(f)
        
        rows = list(reader)
        if not rows:
            return []
            
        header_row_idx = -1
        headers = []
        
        for idx, row in enumerate(rows):
            row_headers = [str(cell).strip().lower() for cell in row]
            
            def has_match(names):
                return any(any(name in cell for name in names) for cell in row_headers)
                
            has_sym = has_match(["symbol", "ticker", "share", "asset", "description", "security", "name"])
            has_qty = has_match(["quantity", "qty", "shares", "units", "quantity sold"])
            has_buy_dt = has_match(["buy date", "purchase date", "buy_date", "purchase_date", "acquired", "date acquired", "acq date", "dt_buy", "opened date", "open date", "opened_date"])
            has_buy_val = has_match(["cost per share", "purchase price per share", "buy price per share", "price per share", "buy price", "purchase price", "cost basis", "buy_val", "cost", "total cost", "purchase value", "purchase rate", "buy rate", "cost basis (cb)", "cost basis (usd)"])
            has_sell_dt = has_match(["sell date", "sale date", "sell_date", "sale_date", "sold", "date sold", "disposal date", "dt_sell", "closed date", "close date", "closed_date", "transaction closed date"])
            has_sell_val = has_match(["proceeds per share", "sale price per share", "sell price per share", "price per share", "sell price", "sale price", "proceeds", "gross proceeds", "total proceeds", "value sold", "sales proceeds", "sale value", "sale rate", "sell rate", "proceeds (usd)"])
            
            matches_count = sum([has_sym, has_qty, has_buy_dt, has_buy_val, has_sell_dt, has_sell_val])
            if matches_count >= 4:
                header_row_idx = idx
                headers = row_headers
                break
                
        if header_row_idx == -1:
            for idx, row in enumerate(rows):
                if any(row):
                    header_row_idx = idx
                    headers = [str(cell).strip().lower() for cell in row]
                    break
                    
        if header_row_idx == -1:
            return []
            
        def get_col_idx(names):
            for name in names:
                for col_idx, h in enumerate(headers):
                    if name == h or name in h:
                        return col_idx
            return -1

        sym_idx = get_col_idx(["symbol", "ticker", "share", "asset", "description", "security", "stock symbol"])
        isin_idx = get_col_idx(["isin", "code"])
        qty_idx = get_col_idx(["quantity sold", "quantity", "qty", "shares", "units"])
        buy_date_idx = get_col_idx(["date acquired", "acquired", "acq date", "opened date", "open date", "opened_date", "buy date", "purchase date", "buy_date", "purchase_date", "dt_buy"])
        buy_price_idx = get_col_idx(["purchase rate", "buy rate", "cost per share", "purchase price per share", "buy price per share", "price per share", "cost basis (cb)", "cost basis (usd)", "cost basis", "cost", "total cost", "purchase price", "buy price", "buy_price", "purchase_price", "purchase value"])
        sell_date_idx = get_col_idx(["date sold", "sold", "closed date", "close date", "closed_date", "transaction closed date", "sell date", "sale date", "sell_date", "sale_date", "disposal date", "dt_sell"])
        sell_price_idx = get_col_idx(["sale rate", "sell rate", "proceeds per share", "sale price per share", "sell price per share", "price per share", "proceeds (usd)", "proceeds", "gross proceeds", "total proceeds", "sell price", "sale price", "sell_price", "sale_price", "value sold", "sales proceeds", "sale value"])
        
        brokerage_idx = get_col_idx(["brokerage", "commission", "commissions", "fee", "fees"])
        stt_idx = get_col_idx(["stt", "securities transaction tax"])
        service_tax_idx = get_col_idx(["service tax", "gst", "cgst", "sgst", "igst"])
        trans_charges_idx = get_col_idx(["transaction charges", "exchange transaction charges", "turnover charges"])
        other_charges_idx = get_col_idx(["other charges", "stamp duty", "stamp_duty", "sebi fees"])

        if any(idx == -1 for idx in [qty_idx, buy_date_idx, buy_price_idx, sell_date_idx, sell_price_idx]) and self.anthropic_client:
            try:
                logger.info(f"Regex column matching failed for some headers. Invoking Claude. Headers: {headers}")
                prompt = f"""
Analyze the following CSV header columns from a stock brokerage statement.
Map them to the required fields:
1. "quantity" (quantity of shares sold)
2. "buy_date" (date the shares were purchased/acquired)
3. "buy_price" (purchase price / cost basis per share or total cost basis)
4. "sell_date" (date the shares were sold)
5. "sell_price" (sale price / proceeds per share or total proceeds)

Return ONLY a valid JSON object matching the headers to these indices.
Available Headers: {headers}
Example Response:
{{"quantity": 2, "buy_date": 3, "buy_price": 5, "sell_date": 4, "sell_price": 6}}
"""
                response = self.anthropic_client.messages.create(
                    model="claude-3-5-haiku-20241022",
                    max_tokens=256,
                    temperature=0.0,
                    messages=[{"role": "user", "content": prompt}]
                )
                map_data = json.loads(response.content[0].text.strip())
                qty_idx = map_data.get("quantity", qty_idx)
                buy_date_idx = map_data.get("buy_date", buy_date_idx)
                buy_price_idx = map_data.get("buy_price", buy_price_idx)
                sell_date_idx = map_data.get("sell_date", sell_date_idx)
                sell_price_idx = map_data.get("sell_price", sell_price_idx)
            except Exception as claude_err:
                logger.error(f"Claude column mapping failed: {claude_err}")

        for row in rows[header_row_idx + 1:]:
            if not row or len(row) <= max([qty_idx, buy_date_idx, buy_price_idx, sell_date_idx, sell_price_idx], default=-1):
                continue
            
            try:
                symbol = row[sym_idx].strip() if sym_idx != -1 else "UNKNOWN"
                isin = row[isin_idx].strip() if isin_idx != -1 else ""
                
                qty = float(row[qty_idx].replace(",", ""))
                if qty <= 0:
                    continue
                
                def parse_date(d_str):
                    d_str = d_str.strip()
                    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y"):
                        try:
                            return datetime.strptime(d_str, fmt).date()
                        except ValueError:
                            continue
                    from dateutil import parser as date_parser
                    return date_parser.parse(d_str).date()

                buy_date = parse_date(row[buy_date_idx])
                sell_date = parse_date(row[sell_date_idx])
                
                raw_buy_val = float(re.sub(r"[^\d\.\-]", "", row[buy_price_idx]))
                raw_sell_val = float(re.sub(r"[^\d\.\-]", "", row[sell_price_idx]))
                
                # If it's a total column (e.g. cost basis, proceeds), divide by quantity to get price per share.
                # If it's already a per-share column, do not divide.
                h_buy = headers[buy_price_idx]
                is_total_cost = "basis" in h_buy or "total" in h_buy or h_buy == "cost"
                buy_price_val = raw_buy_val / qty if (is_total_cost and raw_buy_val > 10 * qty) else raw_buy_val
                
                h_sell = headers[sell_price_idx]
                is_total_proceeds = ("proceeds" in h_sell or "value" in h_sell) and "per share" not in h_sell and "rate" not in h_sell
                sell_price_val = raw_sell_val / qty if (is_total_proceeds and raw_sell_val > 10 * qty) else raw_sell_val
                
                brokerage = float(re.sub(r"[^\d\.\-]", "", row[brokerage_idx])) if brokerage_idx != -1 and row[brokerage_idx] else 0.0
                stt = float(re.sub(r"[^\d\.\-]", "", row[stt_idx])) if stt_idx != -1 and row[stt_idx] else 0.0
                service_tax = float(re.sub(r"[^\d\.\-]", "", row[service_tax_idx])) if service_tax_idx != -1 and row[service_tax_idx] else 0.0
                trans_charges = float(re.sub(r"[^\d\.\-]", "", row[trans_charges_idx])) if trans_charges_idx != -1 and row[trans_charges_idx] else 0.0
                other_charges = float(re.sub(r"[^\d\.\-]", "", row[other_charges_idx])) if other_charges_idx != -1 and row[other_charges_idx] else 0.0
                
                transfer_expenses = brokerage + service_tax + trans_charges + other_charges
                if not is_us:
                    transfer_expenses += stt

                rate_buy_used = 1.0
                rate_sell_used = 1.0
                if is_us:
                    rate_buy_used = self.rate_resolver.resolve_rule_115_rate(buy_date)
                    rate_sell_used = self.rate_resolver.resolve_rule_115_rate(sell_date)
                    
                buy_price_inr = buy_price_val * rate_buy_used
                sell_price_inr = ((sell_price_val * qty - transfer_expenses) * rate_sell_used) / qty
                
                records.append({
                    "symbol": symbol,
                    "isin": isin,
                    "quantity": qty,
                    "buy_date": buy_date,
                    "buy_price": buy_price_val,
                    "buy_price_inr": buy_price_inr,
                    "sell_date": sell_date,
                    "sell_price": sell_price_val,
                    "sell_price_inr": sell_price_inr,
                    "rate_buy_used": rate_buy_used,
                    "rate_sell_used": rate_sell_used,
                    "transfer_expenses": transfer_expenses,
                    "is_us": is_us
                })
            except Exception as e:
                logger.warning(f"Failed to parse row {row}: {e}")
                continue

        return records

    def parse_us_dividends_csv(self, csv_content: str) -> list:
        """Parses US dividends CSV."""
        records = []
        normalized_content = csv_content.replace('\r\n', '\n').replace('\r', '\n')
        f = io.StringIO(normalized_content.strip())
        reader = csv.reader(f)
        
        rows = list(reader)
        if not rows:
            return []

        header_row_idx = -1
        headers = []
        
        for idx, row in enumerate(rows):
            row_headers = [str(cell).strip().lower() for cell in row]
            
            def has_match(names):
                return any(any(name in cell for name in names) for cell in row_headers)
                
            has_dt = has_match(["date", "payment date", "activity date", "transaction date", "dt_pay"])
            has_amt = has_match(["amount", "gross amount", "gross amount (usd)", "net amount", "total", "value"])
            
            if has_dt and has_amt:
                header_row_idx = idx
                headers = row_headers
                break
                
        if header_row_idx == -1:
            for idx, row in enumerate(rows):
                if any(row):
                    header_row_idx = idx
                    headers = [str(cell).strip().lower() for cell in row]
                    break
                    
        if header_row_idx == -1:
            return []

        def get_col_idx(names):
            for name in names:
                for col_idx, h in enumerate(headers):
                    if name == h or name in h:
                        return col_idx
            return -1

        sym_idx = get_col_idx(["symbol", "ticker", "description", "security", "security description"])
        date_idx = get_col_idx(["date", "payment date", "activity date", "transaction date"])
        amt_idx = get_col_idx(["amount", "gross amount", "gross amount (usd)", "value", "total"])
        tax_idx = get_col_idx(["withholding tax", "withholding", "tax", "nra tax", "fed tax"])
        type_idx = get_col_idx(["type", "activity type", "action", "description"])

        is_schwab_format = any("schwab" in str(r).lower() for r in rows[:10]) or any("nra tax adjustment" in str(r).lower() for r in rows)
        
        if is_schwab_format:
            grouped = {}
            for row in rows[header_row_idx + 1:]:
                if not row or len(row) <= max([date_idx, amt_idx, sym_idx], default=-1):
                    continue
                try:
                    action_type = row[type_idx].strip().lower() if type_idx != -1 else ""
                    if not any(k in action_type for k in ["dividend", "div", "nra tax"]):
                        continue
                        
                    date_str = row[date_idx].strip()
                    if not date_str:
                        continue
                        
                    def parse_date(d_str):
                        d_str = d_str.strip()
                        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
                            try:
                                return datetime.strptime(d_str, fmt).date()
                            except ValueError:
                                continue
                        from dateutil import parser as date_parser
                        return date_parser.parse(d_str).date()
                        
                    div_date = parse_date(date_str)
                    amount = float(re.sub(r"[^\d\.\-]", "", row[amt_idx]))
                    symbol = row[sym_idx].strip().upper()
                    
                    key = (symbol, div_date)
                    if key not in grouped:
                        grouped[key] = {"gross": 0.0, "tax": 0.0}
                        
                    if "dividend" in action_type or "div" in action_type:
                        grouped[key]["gross"] += amount
                    elif "nra tax" in action_type:
                        grouped[key]["tax"] += abs(amount)
                except Exception as e:
                    logger.warning(f"Failed to parse Schwab row {row}: {e}")
                    continue
                    
            for (symbol, div_date), vals in grouped.items():
                if vals["gross"] <= 0:
                    continue
                rate = self.rate_resolver.resolve_rule_115_rate(div_date)
                records.append({
                    "symbol": symbol,
                    "date": div_date,
                    "amount_usd": vals["gross"],
                    "amount_inr": vals["gross"] * rate,
                    "withholding_usd": vals["tax"],
                    "withholding_inr": vals["tax"] * rate,
                    "rate_used": rate
                })
        else:
            for row in rows[header_row_idx + 1:]:
                if not row or len(row) <= max([date_idx, amt_idx], default=-1):
                    continue
                try:
                    date_str = row[date_idx].strip()
                    if not date_str:
                        continue
                        
                    def parse_date(d_str):
                        d_str = d_str.strip()
                        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
                            try:
                                return datetime.strptime(d_str, fmt).date()
                            except ValueError:
                                continue
                        from dateutil import parser as date_parser
                        return date_parser.parse(d_str).date()
                        
                    div_date = parse_date(date_str)
                    amount_usd = float(re.sub(r"[^\d\.\-]", "", row[amt_idx]))
                    
                    withholding_usd = 0.0
                    if tax_idx != -1 and len(row) > tax_idx:
                        tax_str = re.sub(r"[^\d\.\-]", "", row[tax_idx])
                        if tax_str:
                            withholding_usd = abs(float(tax_str))
                            
                    symbol = row[sym_idx].strip().upper() if sym_idx != -1 else "UNKNOWN"
                    
                    rate = self.rate_resolver.resolve_rule_115_rate(div_date)
                    records.append({
                        "symbol": symbol,
                        "date": div_date,
                        "amount_usd": amount_usd,
                        "amount_inr": amount_usd * rate,
                        "withholding_usd": withholding_usd,
                        "withholding_inr": withholding_usd * rate,
                        "rate_used": rate
                    })
                except Exception as e:
                    logger.warning(f"Failed to parse column row {row}: {e}")
                    continue

        return records

    def _find_best_value_for_label(self, label_bbox: dict, candidate_words: list, direction: str = "below") -> dict:
        best_candidate = None
        best_score = float('inf')
        
        lx0, lx1, ltop, lbot = label_bbox["x0"], label_bbox["x1"], label_bbox["top"], label_bbox["bottom"]
        
        for w in candidate_words:
            wx0, wx1, wtop, wbot = w["x0"], w["x1"], w["top"], w["bottom"]
            
            if direction == "below":
                is_below = wtop >= lbot + 2 and wtop <= lbot + 35
                horiz_overlap = (wx0 >= lx0 - 5 and wx0 <= lx1 + 5) or (wx1 >= lx0 - 5 and wx1 <= lx1 + 5) or (wx0 <= lx0 and wx1 >= lx1)
                
                if is_below and horiz_overlap:
                    v_dist = wtop - lbot
                    h_shift = abs((wx0 + wx1)/2 - (lx0 + lx1)/2)
                    score = v_dist + h_shift * 0.5
                    if score < best_score:
                        best_score = score
                        best_candidate = w
                        
            elif direction == "right":
                is_right_height = abs(wtop - ltop) <= 10 or abs(wbot - lbot) <= 10
                is_right_side = wx0 >= lx1 - 5 and wx0 <= lx1 + 180
                
                if is_right_height and is_right_side:
                    h_dist = wx0 - lx1
                    v_shift = abs(wtop - ltop)
                    score = h_dist + v_shift * 2
                    if score < best_score:
                        best_score = score
                        best_candidate = w
                        
        return best_candidate

    def parse_1042s(self, pdf_bytes: bytes) -> list:
        """Parses Form 1042-S PDF page-by-page using pdfplumber coordinate-based cells extraction."""
        import pdfplumber
        unique_forms = []
        seen = set()
        
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    text = page.extract_text() or ""
                    if "1042-S" not in text:
                        continue
                        
                    words = page.extract_words()
                    
                    candidate_words = []
                    for w in words:
                        txt = w["text"].strip().replace("..", ".").replace(",", "")
                        if re.match(r"^\d+\.\d{2}$", txt) or re.match(r"^\d+$", txt):
                            candidate_words.append(w)
                            
                    gross_words = [w for w in words if w["text"].lower() in ["gross", "income"]]
                    gross_label_bbox = None
                    for w1 in gross_words:
                        if w1["text"].lower() == "gross":
                            for w2 in gross_words:
                                if w2["text"].lower() == "income" and abs(w2["top"] - w1["top"]) <= 5 and abs(w2["x0"] - w1["x1"]) <= 20:
                                    gross_label_bbox = {
                                        "x0": min(w1["x0"], w2["x0"]),
                                        "x1": max(w1["x1"], w2["x1"]),
                                        "top": min(w1["top"], w2["top"]),
                                        "bottom": max(w1["bottom"], w2["bottom"])
                                    }
                                    break
                            if gross_label_bbox:
                                break
                                
                    withheld_words = [w for w in words if w["text"].lower() in ["federal", "tax", "withheld"]]
                    withheld_bbox = None
                    for w1 in withheld_words:
                        if w1["text"].lower() == "federal":
                            for w2 in withheld_words:
                                if w2["text"].lower() == "tax" and abs(w2["top"] - w1["top"]) <= 5:
                                    for w3 in withheld_words:
                                        if w3["text"].lower() == "withheld" and abs(w3["top"] - w2["top"]) <= 10:
                                            withheld_bbox = {
                                                "x0": min(w1["x0"], w2["x0"], w3["x0"]),
                                                "x1": max(w1["x1"], w2["x1"], w3["x1"]),
                                                "top": min(w1["top"], w2["top"], w3["top"]),
                                                "bottom": max(w1["bottom"], w2["bottom"], w3["bottom"])
                                            }
                                            break
                            if withheld_bbox:
                                break
                                
                    code_bbox = None
                    for w1 in words:
                        if w1["text"].lower() == "income":
                            for w2 in words:
                                if w2["text"].lower() == "code":
                                    on_same_line = abs(w2["top"] - w1["top"]) <= 5 and abs(w2["x0"] - w1["x1"]) <= 25
                                    on_next_line = (w2["top"] > w1["top"] and w2["top"] - w1["top"] <= 12) and abs(w2["x0"] - w1["x0"]) <= 15
                                    if on_same_line or on_next_line:
                                        code_bbox = {
                                            "x0": min(w1["x0"], w2["x0"]),
                                            "x1": max(w1["x1"], w2["x1"]),
                                            "top": min(w1["top"], w2["top"]),
                                            "bottom": max(w1["bottom"], w2["bottom"])
                                        }
                                        break
                            if code_bbox:
                                break
                                
                    tax_year = 2025
                    year_match = re.search(r"Form\s*1042-S\s*\(?(\d{4})\)?", text, re.IGNORECASE)
                    if year_match:
                        tax_year = int(year_match.group(1))
                    else:
                        for w in words:
                            if w["top"] < 150 and re.match(r"^\d{4}$", w["text"]):
                                val = int(w["text"])
                                if 2020 <= val <= 2030:
                                    tax_year = val
                                    break
                                    
                    gross_income_usd = 0.0
                    if gross_label_bbox:
                        w_val = self._find_best_value_for_label(gross_label_bbox, candidate_words, direction="below")
                        if w_val:
                            gross_income_usd = float(w_val["text"].replace("..", ".").replace(",", ""))
                            
                    withholding_tax_usd = 0.0
                    if withheld_bbox:
                        w_val = self._find_best_value_for_label(withheld_bbox, candidate_words, direction="right")
                        if w_val:
                            withholding_tax_usd = float(w_val["text"].replace("..", ".").replace(",", ""))
                            
                    income_code = "06"
                    if code_bbox:
                        w_val = self._find_best_value_for_label(code_bbox, candidate_words, direction="below")
                        if w_val:
                            income_code = w_val["text"].strip().zfill(2)
                            
                    parsed = {
                        "income_code": income_code,
                        "gross_income_usd": gross_income_usd,
                        "withholding_tax_usd": withholding_tax_usd,
                        "tax_year": tax_year,
                        "payment_date": None
                    }
                    
                    rec = (
                        income_code,
                        gross_income_usd,
                        withholding_tax_usd,
                        tax_year
                    )
                    
                    if rec[1] > 0 or rec[2] > 0:
                        if rec not in seen:
                            seen.add(rec)
                            unique_forms.append(parsed)
                            logger.info(f"Parsed Form 1042-S page {page_num+1}: Code={income_code}, Gross=${gross_income_usd:.2f}, Tax=${withholding_tax_usd:.2f}")
                            
        except Exception as e:
            logger.error(f"pdfplumber coordinate 1042-S parser failed: {e}")
            
        return unique_forms

    def parse_us_stock_excel(self, file_bytes: bytes) -> list:
        """Parses US stock sales Excel sheets by converting them to CSV string."""
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        output = io.StringIO()
        writer = csv.writer(output)
        for row in sheet.iter_rows(values_only=True):
            if any(row):
                writer.writerow([str(v) if v is not None else "" for v in row])
        csv_content = output.getvalue()
        return self.parse_stock_sales_csv(csv_content, is_us=True)

    def parse_us_dividends_excel(self, file_bytes: bytes) -> list:
        """Parses US dividends Excel sheets by converting them to CSV string."""
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        output = io.StringIO()
        writer = csv.writer(output)
        for row in sheet.iter_rows(values_only=True):
            if any(row):
                writer.writerow([str(v) if v is not None else "" for v in row])
        csv_content = output.getvalue()
        return self.parse_us_dividends_csv(csv_content)

    def parse_us_stock_pdf(self, file_bytes: bytes) -> list:
        """Parses US Stock PDF statements using pypdf and Claude 3.5 Haiku to extract records."""
        import pypdf
        
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        all_text = ""
        for page in reader.pages[:10]:
            all_text += page.extract_text() or ""
            
        if not all_text.strip():
            return []
            
        prompt = f"""
You are an expert tax parser. Extract all US stock/share sale transaction lots from the following raw text of a US brokerage statement.

Analyze the transactions. Each transaction lot must resolve the corresponding purchase date/price and sell date/price.

Return a JSON list of objects, where each object has exactly these fields:
- "symbol": Ticker symbol or name of the stock (string)
- "quantity": Number of shares sold (float)
- "buy_date": Purchase date in "YYYY-MM-DD" format
- "buy_price_usd": Purchase price per share in USD (float)
- "sell_date": Sale date in "YYYY-MM-DD" format
- "sell_price_usd": Sale price per share in USD (float)
- "transfer_expenses_usd": Total non-STT brokerage/transaction charges in USD for this lot (float, default 0.0)

Raw PDF Text:
{all_text}

JSON Output:
"""
        try:
            response = self.anthropic_client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=2000,
                temperature=0.0,
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )
            content = response.content[0].text.strip()
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()
            raw_records = json.loads(content)
            
            records = []
            for r in raw_records:
                try:
                    buy_date = datetime.strptime(r["buy_date"], "%Y-%m-%d").date()
                    sell_date = datetime.strptime(r["sell_date"], "%Y-%m-%d").date()
                    qty = float(r["quantity"])
                    buy_price_usd = float(r["buy_price_usd"])
                    sell_price_usd = float(r["sell_price_usd"])
                    expenses_usd = float(r.get("transfer_expenses_usd", 0.0))
                    
                    buy_rate = self.rate_resolver.resolve_rule_115_rate(buy_date)
                    sell_rate = self.rate_resolver.resolve_rule_115_rate(sell_date)
                    
                    buy_price_inr = buy_price_usd * buy_rate
                    net_sell_val_usd = (sell_price_usd * qty) - expenses_usd
                    sell_price_inr = (net_sell_val_usd * sell_rate) / qty
                    
                    records.append({
                        "symbol": r["symbol"],
                        "isin": "",
                        "quantity": qty,
                        "buy_date": buy_date,
                        "buy_price": buy_price_usd,
                        "buy_price_inr": buy_price_inr,
                        "sell_date": sell_date,
                        "sell_price": sell_price_usd,
                        "sell_price_inr": sell_price_inr,
                        "rate_buy_used": buy_rate,
                        "rate_sell_used": sell_rate,
                        "transfer_expenses": expenses_usd * sell_rate,
                        "is_us": True
                    })
                except Exception as ex:
                    logger.warning(f"Error parsing US PDF lot {r}: {ex}")
                    continue
            return records
        except Exception as e:
            logger.error(f"Error parsing US stock PDF with Claude: {e}")
            return []
